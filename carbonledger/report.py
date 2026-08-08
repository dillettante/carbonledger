"""집계 리포트 — 배출 레코드 → 조직 탄소발자국(S1+S2+S3) md + xlsx + records.json.

정직성 장치(Phase 0·반증검토 반영):
  · 계수별 출처·신뢰수준을 리포트 표면에 노출(부록 자동생성)
  · Scope 3 미측정 12개 카테고리 + Scope 1·2 미측정 배출원을 명시(부분집계 은폐 금지)
  · 면책 고지를 README가 아니라 리포트 파일 자체에 삽입
  · review 큐(미검증 건)를 '본 수치 미포함'으로 요약
"""
import json
from datetime import date
from pathlib import Path

from . import factors, inventory

_CATS = Path(__file__).parent / "data" / "categories.json"

DISCLAIMER = (
    "본 리포트는 조직 탄소발자국 **추정**이다. 배출권거래제·목표관리제 명세서 등 규제 신고용이 아니다. "
    "거리기반·지출기반 산정은 명세서 방법론과 다르며, 계수 일부는 해외정부공식·학술·사용자입력 등급이다. "
    "신고 전 소관기관(gir.go.kr·한국환경공단)의 확정계수·최신 고시로 재검증할 것."
)

# Scope 1·2에서 본 툴이 자동 산정하지 않는 배출원(부분집계 명시용)
_S12_MISSING = [
    ("Scope 1", "냉매 누출(공조·냉동 HFC, fugitive)", "냉매 충전량·누출률 기반 별도 산정"),
    ("Scope 1", "비상발전기·소각 등 기타 고정연소", "연료 사용량 확보 시 fuel_* 계수로 산정 가능"),
    ("Scope 2", "지역난방 열·스팀", "지사별 열 배출계수(factors.json _reference_only) × 열사용량 수기 산정"),
]


def _load_cats() -> dict:
    return json.loads(_CATS.read_text(encoding="utf-8"))["categories"]


def _sum(records, scope=None, category=None) -> float:
    t = 0.0
    for r in records:
        if scope is not None and r.get("scope") != scope:
            continue
        if category is not None and r.get("category") != category:
            continue
        t += r.get("kgco2e", 0) or 0
    return round(t, 3)


def build(records: list[dict], review_queue: list[dict], out_dir: str,
          period: str | None = None, inv: dict | None = None) -> dict:
    """레코드 → 리포트 3종 생성. 반환: 요약 dict(총량 등)."""
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    s1 = _sum(records, scope=1)
    s2 = _sum(records, scope=2)
    s3 = _sum(records, scope=3)
    total = round(s1 + s2 + s3, 3)

    # records.json — 감사추적 원장 (어느 툴·계수판으로 산정했는지 스탬프)
    from . import __version__
    (out / "records.json").write_text(
        json.dumps({"period": period, "generated": str(date.today()),
                    "tool_version": __version__,
                    "factors_version": factors.meta().get("version", ""),
                    "inventory": inv,
                    "records": records, "review_queue": review_queue},
                   ensure_ascii=False, indent=2), encoding="utf-8")

    _write_md(out / "report.md", records, review_queue, s1, s2, s3, total, period, inv)
    _write_xlsx(out / "report.xlsx", records, review_queue, s1, s2, s3, total, period, inv)

    return {"scope1": s1, "scope2": s2, "scope3": s3, "total_kgco2e": total,
            "records": len(records), "review": len(review_queue)}


def _t(kg: float) -> str:
    return f"{kg/1000:,.3f} tCO2eq ({kg:,.1f} kg)"


def _pct(part: float, whole: float) -> str:
    return f"{part / whole * 100:.2f}%" if whole else "—"


def _write_md(path, records, review_queue, s1, s2, s3, total, period, inv=None):
    L = []
    L.append(f"# 조직 온실가스 배출량 리포트")
    L.append(f"\n- 보고기간: **{period or '전체(미지정)'}**  · 생성일: {date.today()}")
    L.append(f"- 자동 산정 건수: {len(records)}  · 검토 대기(미포함): {len(review_queue)}\n")

    L += inventory.render_md(inv)

    L.append("## 1. 총괄 — 조직 온실가스 배출량\n")
    L.append("| 구분 | 배출량 | 비중 |")
    L.append("|---|---|---|")
    L.append(f"| Scope 1 (직접) | {_t(s1)} | {_pct(s1, total)} |")
    L.append(f"| Scope 2 (전력, location-based) | {_t(s2)} | {_pct(s2, total)} |")
    # 실무 데이터표(예: 지속가능경영보고서 ESG 데이터)의 굵은 '온실가스 배출량' 행은
    # 예외 없이 Scope 1+2다 — 규제 대상·감축목표·원단위 분자가 모두 S1+2이기 때문.
    # S1+2+3만 굵게 내면 Scope 3가 지배적일 때 헤드라인이 오도된다.
    L.append(f"| **소계 (Scope 1+2)** | **{_t(s1 + s2)}** | {_pct(s1 + s2, total)} |")
    L.append(f"| Scope 3 (기타 간접) | {_t(s3)} | {_pct(s3, total)} |")
    L.append(f"| **총계 (Scope 1+2+3)** | **{_t(total)}** | 100% |")
    L.append("\n> **소계(Scope 1+2)** 는 규제·감축목표·원단위 산정의 통상 기준이고, "
             "**총계(Scope 1+2+3)** 는 밸류체인 전체 발자국이다. 용도에 맞는 수치를 인용할 것.")
    L.append("> Scope 2는 location-based 단일 산정이다. market-based(녹색프리미엄·REC·PPA) 미반영.")
    L.append("> Scope 1 연료는 **CO2 단독** 산정(CH4·N2O 미가산, 통상 <3%) — 합계는 계수별 GWP 기준이 혼재된 추정치다(부록 §5 참조).\n")

    L.append("## 2. Scope 3 카테고리별 (GHG Protocol 15개 프레임)\n")
    L.append("| # | 카테고리 | 상태 | 배출량 | 산정방법 |")
    L.append("|---|---|---|---|---|")
    cats = _load_cats()
    s3_total = _sum(records, scope=3)
    for i in range(1, 16):
        c = cats[str(i)]
        has = any(r.get("scope") == 3 and r.get("category") == i for r in records)
        if has:
            v = _sum(records, scope=3, category=i)
            # Scope 3 내 비중 5% 초과는 SBTi 넷제로 표준 2.0의 목표 포함 기준선이자
            # ISO 14064-1 5.2.3 '유의한 간접배출' 판단의 실무 참고점 — 눈에 띄게 표시한다.
            share = (v / s3_total * 100) if s3_total else 0
            flag = " **▲5%↑**" if share > 5 else ""
            status = f"측정 ({share:.1f}%{flag})"
            val = _t(v)
            method = str(c.get("method", "")).replace("|", "\\|")
        else:
            # 입력 없는 카테고리는 자동화 여부 무관하게 '미측정' + 측정법 안내(0을 측정으로 오독 방지)
            who = c.get("applies_to", "")
            status = "미측정"
            val = "— (입력 없음)" + (f" · 해당: {who[:20]}" if who else "")
            method = str(c.get("method", c.get("guidance", ""))).replace("|", "\\|")
        L.append(f"| {i} | {c['name']} | {status} | {val} | {method} |")

    # 산정 카테고리 명세 — 실무 공시가 각주로 "12개 카테고리: 1,2,…"를 밝히는 것과 같은 취지.
    # 단 '제외'인지 '미수집'인지는 조직 판단이라 툴이 단정하지 않는다.
    measured = [i for i in range(1, 16)
                if any(r.get("scope") == 3 and r.get("category") == i for r in records)]
    ms = ", ".join(map(str, measured)) or "없음"
    L.append(f"\n> **산정 카테고리: 15개 중 {len(measured)}개** ({ms}). "
             "나머지는 입력이 없어 산정되지 않았을 뿐이며, **해당 없음(제외)인지 미수집인지는 "
             "이 툴이 판정하지 않는다** — 조직이 제외 사유를 별도 기재해야 한다"
             "(ISO 14064-1 5.2.3·9.3.1 i / SBTi 2.0은 제외의 정량 근거를 요구).\n")

    L.append("\n## 3. Scope 1·2 미측정 배출원 (부분집계 고지)\n")
    L.append("아래는 본 툴이 자동 산정하지 않는다. 헤드라인 합계는 이 항목을 **제외**한 부분집계다.\n")
    L.append("| Scope | 배출원 | 측정법 |")
    L.append("|---|---|---|")
    for sc, src, how in _S12_MISSING:
        L.append(f"| {sc} | {src} | {how} |")

    L.append("\n## 4. 건별 명세 (감사추적)\n")
    L.append("| 파일 | Scope | 활동 | 활동량 | 계수 | 계수 출처 | 배출량(kg) |")
    L.append("|---|---|---|---|---|---|---|")
    for r in records:
        # 사용자 입력 계수(user_factor·pcaf_financed)는 레지스트리에 출처가 없으므로
        # 행 자체가 출처를 들고 다닌다. 이 열이 없으면 §5 부록의 "행별 출처 참조" 안내가
        # 가리킬 곳이 없어진다(죽은 포인터).
        src = str(r.get("factor_source", "") or "").replace("|", "\\|")
        L.append(f"| {r.get('source_file','')} | S{r.get('scope','')} | "
                 f"{r.get('activity','')} | {r.get('activity_value','')} {r.get('activity_unit','')} | "
                 f"`{r.get('factor_id','')}` | {src} | {r.get('kgco2e','')} |")

    # 거리 산정 근거 — 지명이 어느 좌표로 해석됐는지. 거리(km)만 남기면
    # 지도 API 1순위가 동명이지·유사상호로 엉뚱한 곳이어도 감사에서 보이지 않는다.
    geo_rows = [r for r in records if r.get("geocoding", {}).get("origin_resolved")]
    if geo_rows:
        L.append("\n### 4-1. 거리 산정 근거 (출장 — 지명→좌표 해석)\n")
        L.append("| 파일 | 출발(해석 결과) | 도착(해석 결과) | 대권거리 | 우회계수 | 적용거리 |")
        L.append("|---|---|---|---|---|---|")
        for r in geo_rows:
            g = r["geocoding"]
            o, d = g["origin_resolved"], g["destination_resolved"]
            det = g.get("detour_factor", 1.0)
            L.append(
                f"| {r.get('source_file','')} | {o['query']} → {o['source']} "
                f"({o['lat']:.4f}, {o['lon']:.4f}) | {d['query']} → {d['source']} "
                f"({d['lat']:.4f}, {d['lon']:.4f}) | {g.get('great_circle_km','')} km | "
                f"×{det} | {r.get('activity_value','')} km |")
        L.append("\n> 좌표 출처가 지도 API(Kakao·Naver)면 **검색 1순위 결과**다 — 동명 지점·"
                 "유사 상호로 엉뚱한 곳이 잡힐 수 있으니 해석 결과를 확인할 것. "
                 "거리는 대권(직선) 근사이며 철도·버스는 우회계수로 실노선을 근사한다.\n")

    _n = [5]  # 이후 절은 조건부 출력이라 번호를 동적으로 매긴다(하드코딩 시 §5→§7 드리프트)
    L.append("\n## 5. 사용된 배출계수 · 출처 (부록)\n")
    used = factors.all_used(r.get("factor_id") for r in records)

    # 계수 신뢰수준 구성 — 헤드라인의 몇 %가 권위 계수이고 몇 %가 사용자 입력인가.
    # 검증인·독자가 가장 먼저 묻는 질문이라 표 앞에 배치한다.
    by_conf = {}
    for r in records:
        fid = r.get("factor_id")
        conf = next((f.get("confidence", "미상") for f in used if f["id"] == fid), "미상")
        by_conf[conf] = by_conf.get(conf, 0) + (r.get("kgco2e", 0) or 0)
    if by_conf and total:
        parts = " · ".join(f"{k} {v/total*100:.1f}%"
                           for k, v in sorted(by_conf.items(), key=lambda x: -x[1]))
        L.append(f"**계수 신뢰수준 구성(배출량 기준)**: {parts}\n")

    L.append("| factor_id | 값 | 단위 | 신뢰수준 | 연도 | GWP | 출처 | 비고(한계·누락) |")
    L.append("|---|---|---|---|---|---|---|---|")
    for f in used:
        # 절단 금지 — 출처의 공표시점·고시번호와 비고의 한계가 잘려나가면
        # 이 부록의 존재 이유(추적 가능성)가 사라진다. 표 파이프만 이스케이프.
        src = str(f.get("source", "")).replace("|", "\\|")
        note = str(f.get("note", "")).replace("|", "\\|")
        L.append(f"| `{f['id']}` | {f.get('value')} | {f.get('unit')} | "
                 f"{f.get('confidence')} | {f.get('year','')} | {f.get('gwp_basis','')} | "
                 f"{src} | {note} |")
    L.append("\n> 비고의 '한계·누락'을 확인할 것. 예: 연료계수는 **CO2만 반영**(CH4·N2O 별도 가산 필요), "
             "전력 WTT/T&D는 UK 프록시 등 — 헤드라인 수치에 영향. "
             "(전력계수 0.4173은 gir 원문 검증 완료 — GWP=AR5.)")
    L.append("> `user_factor`·`pcaf_financed`는 레지스트리 계수가 아니라 **사용자가 입력한 계수**다 — "
             "행별 출처는 §4 건별 명세의 '계수 출처' 열 참조.")

    # 수기 교정 이력 — 자동 추출 건과 구별되지 않으면 통제 흔적이 사라진다
    corrected = [r for r in records if r.get("human_corrected")]
    if corrected:
        _n[0] += 1
        L.append(f"\n## {_n[0]}. 수기 교정 이력 (사람이 값을 확인·수정해 합계에 반영한 건)\n")
        L.append(f"아래 {len(corrected)}건은 자동 추출이 아니라 **사람이 교정**해 집계에 포함됐다. "
                 "교정본도 검증 관문(필수필드·산술 일치·교정 이력)을 통과한 것만 반영된다.\n")
        L.append("| 파일 | 활동 | 배출량(kg) | 교정자 | 교정일시 | 근거 |")
        L.append("|---|---|---|---|---|---|")
        for r in corrected:
            rv = r.get("review") or {}
            L.append(f"| {r.get('source_file','')} | {r.get('activity','')} | {r.get('kgco2e','')} | "
                     f"{rv.get('reviewer','')} | {rv.get('reviewed_at','')} | {rv.get('basis','')} |")
        L.append(f"\n> 수기 교정분 합계: {_t(sum(r.get('kgco2e', 0) or 0 for r in corrected))} "
                 f"(전체의 {(sum(r.get('kgco2e',0) or 0 for r in corrected) / total * 100 if total else 0):.1f}%)\n")

    if review_queue:
        _n[0] += 1
        L.append(f"\n## {_n[0]}. 검토 대기 (본 수치 미포함)\n")
        L.append("검증 관문을 통과 못해 집계에서 제외됐다. 교정 후 `carbonledger review`로 재집계.\n")
        L.append("| 파일 | 사유 |")
        L.append("|---|---|")
        for q in review_queue:
            L.append(f"| {q.get('source_file','')} | {'; '.join(q.get('issues',[]))} |")

    L.append(f"\n---\n\n> ⚠️ **면책** — {DISCLAIMER}\n")
    Path(path).write_text("\n".join(L), encoding="utf-8")


def _write_xlsx(path, records, review_queue, s1, s2, s3, total, period, inv=None):
    from openpyxl import Workbook
    wb = Workbook()

    ws = wb.active
    ws.title = "총괄"
    ws.append(["보고기간", period or "전체(미지정)"])
    ws.append(["생성일", str(date.today())])
    ws.append([])
    ws.append(["구분", "배출량(kgCO2eq)", "tCO2eq"])
    # md §1과 동일하게 소계(S1+2)를 낸다 — 규제·감축목표·원단위의 통상 기준이라
    # xlsx만 유통돼도 헤드라인이 오도되지 않게 한다
    for name, v in [("Scope 1", s1), ("Scope 2", s2), ("소계 (Scope 1+2)", round(s1 + s2, 3)),
                    ("Scope 3", s3), ("총계 (Scope 1+2+3)", total)]:
        ws.append([name, v, round(v / 1000, 3)])
    ws.append([])
    ws.append(["면책", DISCLAIMER])

    # 조직 선언 — md §0과 대칭. xlsx 단독 유통 시 "무엇에 대한 숫자인지"가 소실되지 않게
    wi = wb.create_sheet("조직선언")
    if inv:
        wi.append(["항목", "내용"])
        for key, label, _ in inventory.FIELDS:
            wi.append([label, str(inv.get(key, "") or "") or "(미기재)"])
        by = inv.get("base_year") or {}
        if isinstance(by, dict) and (by.get("year") or by.get("note")):
            wi.append(["기준연도", f"{by.get('year', '')} {by.get('note', '') or ''}".strip()])
        for e in inv.get("exclusions") or []:
            wi.append([f"제외: {e.get('item', '')}", e.get("reason", "") or "(사유 미기재)"])
        v = inv.get("verification") or {}
        if v.get("status"):
            wi.append(["검증(조직 선언)", " · ".join(
                str(v.get(k, "") or "") for k in ("status", "body", "level", "date") if v.get(k))])
            wi.append(["주의", "위 검증은 조직 인벤토리에 관한 선언이며 본 산정치의 검증이 아님"])
        wi.append(["주의", "조직이 기재한 선언의 전재이며 툴은 적정성을 판정하지 않음"])
    else:
        wi.append(["조직 선언", "없음 — input/inventory.json 미제공(배출량 산정 결과만 수록)"])

    wr = wb.create_sheet("건별_감사추적")
    wr.append(["파일", "Scope", "카테고리", "활동", "활동량", "활동단위",
               "factor_id", "계수값", "계수단위", "계수출처", "배출량(kg)",
               "수기교정", "교정자", "교정일시", "교정근거", "거리산정근거"])
    for r in records:
        rv = r.get("review") or {}
        g = r.get("geocoding") or {}
        geo = ""
        if g.get("origin_resolved"):
            o, d = g["origin_resolved"], g["destination_resolved"]
            geo = (f"{o['query']}→{o['source']} ({o['lat']:.4f},{o['lon']:.4f}) / "
                   f"{d['query']}→{d['source']} ({d['lat']:.4f},{d['lon']:.4f}) / "
                   f"대권 {g.get('great_circle_km')}km ×{g.get('detour_factor')}")
        wr.append([r.get("source_file"), r.get("scope"), r.get("category"),
                   r.get("activity"), r.get("activity_value"), r.get("activity_unit"),
                   r.get("factor_id"), r.get("factor_value"), r.get("factor_unit"),
                   r.get("factor_source", ""), r.get("kgco2e"),
                   "Y" if r.get("human_corrected") else "",
                   rv.get("reviewer", ""), rv.get("reviewed_at", ""), rv.get("basis", ""),
                   geo])

    wf = wb.create_sheet("계수목록")
    wf.append(["factor_id", "값", "단위", "신뢰수준", "연도", "GWP기준", "출처", "출처URL"])
    for f in factors.all_used(r.get("factor_id") for r in records):
        wf.append([f["id"], f.get("value"), f.get("unit"), f.get("confidence"),
                   f.get("year"), f.get("gwp_basis"), f.get("source"), f.get("source_url")])

    if review_queue:
        wq = wb.create_sheet("검토대기")
        wq.append(["파일", "사유"])
        for q in review_queue:
            wq.append([q.get("source_file"), "; ".join(q.get("issues", []))])

    # 수식 주입 봉쇄 — openpyxl은 '='로 시작하는 문자열을 살아있는 수식으로 저장한다.
    # 파일명·LLM 추출값이 셀에 그대로 들어가므로, 이 리포트의 모든 셀은 '값'으로 강제한다.
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    cell.data_type = "s"

    wb.save(path)


def selftest():
    import tempfile
    recs = [
        {"source_file": "a.jpg", "scope": 1, "activity": "경유 연소",
         "activity_value": 100, "activity_unit": "L", "factor_id": "fuel_diesel",
         "factor_value": 2.577, "factor_unit": "kgCO2/L", "kgco2e": 257.7},
        {"source_file": "b.jpg", "scope": 3, "category": 6, "activity": "KTX 출장",
         "activity_value": 400, "activity_unit": "passenger-km", "factor_id": "travel_rail_ktx",
         "factor_value": 0.0269, "factor_unit": "kgCO2eq/passenger-km", "kgco2e": 10.76},
    ]
    q = [{"source_file": "c.pdf", "issues": ["역명 확인 필요: '서을'"]}]
    with tempfile.TemporaryDirectory() as d:
        summary = build(recs, q, d, period="2026")
        assert summary["scope1"] == 257.7 and summary["scope3"] == 10.76, "집계 오류"
        assert summary["total_kgco2e"] == round(257.7 + 10.76, 3), "합계 오류"
        md = (Path(d) / "report.md").read_text(encoding="utf-8")
        assert "면책" in md and "미측정" in md and "travel_rail_ktx" in md, "리포트 정직성장치 누락"
        assert (Path(d) / "report.xlsx").exists(), "xlsx 미생성"
        assert (Path(d) / "records.json").exists(), "records.json 미생성"
    print("report selftest 통과 ✅")


if __name__ == "__main__":
    selftest()
