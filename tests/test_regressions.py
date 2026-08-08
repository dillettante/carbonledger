"""회귀 테스트 — 2026-08 전면 재검토(critic)에서 실증 재현된 결함을 못박는다.

각 테스트는 '재현 → 수정 확인' 순서로 만들어졌다. 여기 있는 케이스는 전부
실제로 크래시하거나 틀린 숫자를 냈던 입력이다(가상의 방어가 아니다).
LLM 호출은 골든 테스트와 같은 방식으로 목킹한다 — 네트워크 불필요, CI 가능.
"""
import json
import sys
from pathlib import Path
from types import SimpleNamespace

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from carbonledger import calc, cli, extract, report, scope3, validate  # noqa: E402


# ── 크래시 1: 계수 미등재 수단('기타'·택시)이 실행 전체를 죽이던 결함 ──
def test_transport_unknown_mode_gated(tmp_path):
    """프롬프트가 허용하는 '기타'가 validate를 통과해 calc._TRANSPORT KeyError로
    run 전체가 죽었다. 이제 validate가 계수 미등재 수단을 큐로 보낸다."""
    rec = {"transport": "기타", "origin": "서울시청", "destination": "부산시청",
           "date": "2026-07-12", "amount": 15000}
    assert any("산정 불가 수단" in i for i in validate.validate_transport(rec)), \
        "'기타' 수단이 검증을 통과(KeyError 크래시 경로 재개방)"

    # 화이트리스트와 계수표가 어긋나면 같은 크래시가 재발한다 — 동기화 불변식
    assert set(validate._TRANSPORT_ALLOW) == set(calc._TRANSPORT), \
        "validate 화이트리스트와 calc._TRANSPORT 수단 집합 불일치"


# ── 크래시 2: 가스 고지서 unit=null → AttributeError ──
def test_gas_unit_null_queued(tmp_path, monkeypatch):
    """LLM은 지침대로 unit을 null로 줄 수 있다. rec.get('unit','')는 값 null에
    무력해 .replace에서 죽었다 — 이제 큐 사유가 나와야 한다."""
    issues = validate.validate_gas({"usage": 50, "unit": None, "amount": 30000,
                                    "billing_month": "2026-05"})
    assert any("단위" in i for i in issues), "unit=null이 사유 없이 통과"

    # 파이프라인 통합: 추출이 unit=null을 줘도 run이 죽지 않고 큐로
    (tmp_path / "scope1-fuel").mkdir()
    (tmp_path / "scope1-fuel" / "도시가스_202605.png").write_bytes(b"\x89PNG fake")
    monkeypatch.setattr(extract, "extract", lambda p, d, m=None: {
        "doc": "도시가스", "usage": 50, "unit": None, "amount": 30000,
        "billing_month": "2026-05"})
    records, queue = [], []
    cli._process_bills(tmp_path / "scope1-fuel", "scope1-fuel", None, "2026", records, queue)
    assert records == [] and len(queue) == 1, "unit=null 건이 큐로 가지 않음"


# ── 크래시 3: 지오코딩 네트워크·쿼터 예외가 배치 전체를 죽이던 결함 ──
def test_geocode_error_queued(tmp_path, monkeypatch):
    import requests

    def boom(place):
        raise requests.ConnectionError("모의 쿼터 소진")
    monkeypatch.setattr(calc, "_api_geocode", boom)

    records, queue = [], []
    rec = {"transport": "버스", "origin": "서울고속터미널", "destination": "부산",
           "date": "2026-07-12", "amount": 30000}
    cli._calc_travel("버스표.png", "transport", rec, "2026", records, queue)  # 크래시하면 실패
    assert records == [] and len(queue) == 1
    assert any("거리 산정 오류" in i for i in queue[0]["issues"])


# ── 정합성 1: review 병합이 cat3 파생을 재계산하지 않던 결함 ──
def test_review_rederives_cat3(tmp_path):
    """전기 1000→2000kWh 교정 시: 종전엔 1000kWh 기반 파생(WTT·T&D)이 스테일로
    남았다. 이제 병합 후 파생을 전부 다시 만든다."""
    out = tmp_path / "out"
    records = [{"source_file": "전기_202605.png", "scope": 2, "category": None,
                "activity": "전력 사용", "factor_id": "electricity_kr",
                "factor_value": 0.4173, "factor_unit": "kgCO2eq/kWh",
                "activity_value": 1000, "activity_unit": "kWh", "kgco2e": 417.3}]
    scope3.derive_category3(records, [])
    assert len(records) == 3, "전제: 전력 1건에서 파생 2건(WTT·T&D)"
    report.build(records, [], str(out), period="2026")

    rev = out / "reviewed"
    rev.mkdir()
    (rev / "fix1.json").write_text(json.dumps({
        "source_file": "전기_202605.png", "scope": 2, "activity": "전력 사용",
        "factor_id": "electricity_kr", "factor_value": 0.4173,
        "factor_unit": "kgCO2eq/kWh", "activity_value": 2000,
        "activity_unit": "kWh", "kgco2e": 834.6,
        "review": {"reviewer": "홍길동", "reviewed_at": "2026-08-09",
                   "basis": "고지서 재확인 — 2,000kWh"}}, ensure_ascii=False),
        encoding="utf-8")
    cli.cmd_review(SimpleNamespace(out=str(out)))

    data = json.loads((out / "records.json").read_text(encoding="utf-8"))
    cat3 = [r for r in data["records"] if r.get("category") == 3]
    assert len(cat3) == 2, f"파생 개수 이상: {len(cat3)}"
    assert all(r["activity_value"] == 2000 for r in cat3), \
        f"스테일 파생 잔존: {[r['activity_value'] for r in cat3]}"


# ── 정합성 2: source_file 오타 교정본이 유령 레코드로 총계 진입하던 결함 ──
def test_review_rejects_ghost_and_derived(tmp_path):
    out = tmp_path / "out"
    records = [{"source_file": "a.png", "scope": 1, "activity": "경유",
                "factor_id": "fuel_diesel", "factor_value": 2.577,
                "activity_value": 100, "activity_unit": "L", "kgco2e": 257.7}]
    report.build(records, [], str(out), period="2026")
    rev = out / "reviewed"
    rev.mkdir()
    hist = {"review": {"reviewer": "홍길동", "reviewed_at": "2026-08-09", "basis": "확인"}}
    # ①대상 없는 source_file(오타) ②자동 파생 건 직접 교정 — 둘 다 반려돼야
    (rev / "ghost.json").write_text(json.dumps({
        "source_file": "없는파일.png", "scope": 1, "factor_id": "fuel_diesel",
        "factor_value": 2.577, "activity_value": 10, "activity_unit": "L",
        "kgco2e": 25.77, **hist}, ensure_ascii=False), encoding="utf-8")
    (rev / "derived.json").write_text(json.dumps({
        "source_file": "a.png→cat3", "scope": 3, "category": 3,
        "factor_value": 0.61101, "activity_value": 100, "kgco2e": 61.101,
        **hist}, ensure_ascii=False), encoding="utf-8")
    cli.cmd_review(SimpleNamespace(out=str(out)))

    data = json.loads((out / "records.json").read_text(encoding="utf-8"))
    srcs = [r["source_file"] for r in data["records"]]
    assert "없는파일.png" not in srcs, "유령 교정본이 총계에 진입"
    issues = "; ".join(i for q in data["review_queue"] for i in q.get("issues", []))
    assert "교정 대상 없음" in issues, "유령 반려 사유 없음"
    assert "직접 교정 불가" in issues, "파생 직접 교정이 반려되지 않음"


# ── 정합성 3: spend·commute CSV에 fail-closed 관문이 없던 결함 ──
def test_spend_commute_gates(tmp_path):
    p = tmp_path / "spend.csv"
    p.write_text("item,krw,factor,factor_source\n환불오타,-50000000,0.0004,한국은행표\n",
                 encoding="utf-8")
    records, queue = [], []
    cli._process_spend(p, records, queue)
    assert records == [] and len(queue) == 1, "음수 지출이 총계를 조용히 깎음"

    p2 = tmp_path / "commute.csv"
    p2.write_text("employee_id,mode,factor_id,oneway_km,workdays\n"
                  "E001,지하철,commute_subway,18,2200\n"
                  "E002,지하철,commute_subway,-5,220\n", encoding="utf-8")
    records2, queue2 = [], []
    cli._process_commute(p2, records2, queue2)
    assert records2 == [] and len(queue2) == 2, "비상식 통근값(근무일 2200·음수 거리) 통과"


# ── 정합성 4: 통근 factor_id에 아무 계수나 넣어도 통과하던 결함 ──
def test_commute_factor_guard(tmp_path):
    """단위검사가 계수 분모를 활동단위로 삼아 자기충족이었다 — electricity_kr을
    넣으면 km가 kWh로 둔갑해 무경고 오산정됐다."""
    try:
        calc.scope3_commute("electricity_kr", 12000)
        assert False, "통근이 아닌 계수를 통과시킴"
    except Exception as e:
        assert "통근 계수가 아님" in str(e)


# ── 정합성 5: PCAF 귀속계수 > 1 무경고 ──
def test_pcaf_attribution_over_1(tmp_path):
    p = tmp_path / "cat15_inv.csv"
    p.write_text("asset,asset_class,outstanding,denominator,investee_emissions,emissions_source\n"
                 "A사,상장주식,1200000000,1000000000,500,공시\n", encoding="utf-8")
    records, queue = [], []
    scope3.process_pcaf(p, records, queue)
    assert records == [] and len(queue) == 1, "귀속>1(잔액>기업가치)이 무경고 통과"
    assert "귀속계수" in queue[0]["issues"][0]


# ── 보안: xlsx 수식 주입('='로 시작하는 셀이 수식으로 저장) ──
def test_xlsx_no_formula_injection(tmp_path):
    from openpyxl import load_workbook
    recs = [{"source_file": '=HYPERLINK("http://evil.example","x")',
             "scope": 1, "activity": "경유", "activity_value": 100,
             "activity_unit": "L", "factor_id": "fuel_diesel",
             "factor_value": 2.577, "factor_unit": "kgCO2/L", "kgco2e": 257.7}]
    report.build(recs, [], str(tmp_path), period="2026")
    wb = load_workbook(tmp_path / "report.xlsx")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                assert cell.data_type != "f", \
                    f"수식 셀 발견: {ws.title}!{cell.coordinate} = {cell.value!r}"
    # 원문은 보존돼야 한다(이스케이프로 값이 훼손되면 감사추적 실패)
    a2 = wb["건별_감사추적"]["A2"]
    assert a2.value.startswith("=HYPERLINK"), "셀 값 원문 훼손"


# ── 교정본: Scope 3에 category 없으면 §2 표합과 총계가 어긋나던 결함 ──
def test_scope3_correction_requires_category(tmp_path):
    base = {"source_file": "x.csv#1", "scope": 3, "factor_value": 1.0,
            "activity_value": 10, "kgco2e": 10.0,
            "review": {"reviewer": "홍길동", "reviewed_at": "2026-08-09", "basis": "확인"}}
    assert any("category" in i for i in validate.validate_corrected_record(base)), \
        "Scope 3 교정본이 category 없이 통과"
    assert validate.validate_corrected_record({**base, "category": 5}) == [], \
        "category 있는 정상 교정본이 반려됨"


# ── 완결성: scope3/의 비대상 CSV가 stdout에만 남고 리포트에 안 남던 결함 ──
def test_catn_skip_to_queue(tmp_path):
    (tmp_path / "scope3").mkdir()
    (tmp_path / "scope3" / "cat6_출장.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "scope3" / "catalog.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    a = SimpleNamespace(input=str(tmp_path), period=None, model=None,
                        out=str(tmp_path / "out"))
    cli.cmd_run(a)
    data = json.loads((tmp_path / "out" / "records.json").read_text(encoding="utf-8"))
    srcs = [q["source_file"] for q in data["review_queue"]]
    assert "cat6_출장.csv" in srcs, "비대상 카테고리 CSV가 리포트에 안 남음"
    assert "catalog.csv" in srcs, "번호 없는 cat*.csv가 조용히 무시됨"


# ── xlsx: 소계·조직선언 시트(md와의 정보 비대칭 해소) ──
def test_xlsx_subtotal_and_inventory_sheet(tmp_path):
    from openpyxl import load_workbook
    recs = [{"source_file": "a.jpg", "scope": 1, "activity": "경유",
             "activity_value": 100, "activity_unit": "L", "factor_id": "fuel_diesel",
             "factor_value": 2.577, "factor_unit": "kgCO2/L", "kgco2e": 257.7}]
    inv = {"organization": "(주)테스트", "consolidation_approach": "운영통제"}
    report.build(recs, [], str(tmp_path), period="2026", inv=inv)
    wb = load_workbook(tmp_path / "report.xlsx")
    assert "조직선언" in wb.sheetnames, "조직선언 시트 없음"
    vals = [c.value for row in wb["총괄"].iter_rows() for c in row if c.value]
    assert any("소계" in str(v) for v in vals), "총괄 시트에 S1+2 소계 없음"
    inv_vals = [c.value for row in wb["조직선언"].iter_rows() for c in row if c.value]
    assert any("(주)테스트" in str(v) for v in inv_vals), "조직 선언 미전재"


# ── 감사추적: 지명→좌표 해석 근거가 레코드·리포트에 남는가 ──
def test_geocoding_audit_trail(tmp_path, monkeypatch):
    """거리(km)만 남기면 지도 API 오매칭(동명이지)을 감사에서 잡을 수 없다."""
    calls = []

    def fake_api(place):
        calls.append(place)
        calc._note_match(place, f"{place} 1순위매칭결과")
        return (37.5, 127.0) if "서울" in place else (35.1, 129.0)

    monkeypatch.setattr(calc, "_api_geocode", fake_api)
    calc._GEOCODE_CACHE.clear()

    records, queue = [], []
    rec = {"transport": "버스", "origin": "서울터미널", "destination": "부산터미널",
           "date": "2026-07-12", "amount": 30000}
    cli._calc_travel("버스표.png", "transport", rec, "2026", records, queue)
    assert len(records) == 1, f"산정 실패: {queue}"

    g = records[0]["geocoding"]
    assert g["origin_resolved"]["lat"] == 37.5, "해석 좌표 미기록"
    assert "1순위매칭결과" in g["origin_resolved"]["source"], \
        f"매칭된 장소명이 안 남음 — 오매칭 감사 불가: {g}"
    assert g["detour_factor"] == 1.2 and g["great_circle_km"] > 0

    # 리포트(md·xlsx) 양쪽에 드러나야 — records.json에만 있으면 사람이 못 본다
    from openpyxl import load_workbook
    report.build(records, [], str(tmp_path), period="2026")
    md = (tmp_path / "report.md").read_text(encoding="utf-8")
    assert "거리 산정 근거" in md and "1순위매칭결과" in md, "md에 산정 근거 누락"
    ws = load_workbook(tmp_path / "report.xlsx")["건별_감사추적"]
    assert ws.cell(1, 16).value == "거리산정근거", "xlsx 감사추적 열 누락"
    assert "1순위매칭결과" in str(ws.cell(2, 16).value), "xlsx에 산정 근거 미기록"

    # 캐시: 같은 지명 반복 조회가 API 쿼터를 태우지 않아야
    before = len(calls)
    cli._calc_travel("버스표2.png", "transport", dict(rec), "2026", records, queue)
    assert len(calls) == before, f"동일 지명 재조회 발생({len(calls) - before}회)"
    calc._GEOCODE_CACHE.clear()


if __name__ == "__main__":
    import tempfile

    class _P:  # monkeypatch 대체(순수 실행용)
        def setattr(self, obj, name, val):
            setattr(obj, name, val)

    for fn, needs_mp in [(test_transport_unknown_mode_gated, False),
                         (test_gas_unit_null_queued, True),
                         (test_geocode_error_queued, True),
                         (test_review_rederives_cat3, False),
                         (test_review_rejects_ghost_and_derived, False),
                         (test_spend_commute_gates, False),
                         (test_commute_factor_guard, False),
                         (test_pcaf_attribution_over_1, False),
                         (test_xlsx_no_formula_injection, False),
                         (test_scope3_correction_requires_category, False),
                         (test_catn_skip_to_queue, False),
                         (test_xlsx_subtotal_and_inventory_sheet, False),
                         (test_geocoding_audit_trail, True)]:
        with tempfile.TemporaryDirectory() as d:
            fn(Path(d), _P()) if needs_mp else fn(Path(d))
    print("회귀 테스트 13종 통과 ✅")
